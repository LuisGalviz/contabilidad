"""
Genera archivos Excel de prueba para los 3 módulos de informes.
Uso: python scripts/generate_test_files.py
Los archivos se crean en scripts/test_files/
"""
from __future__ import annotations

import random
from datetime import date, timedelta
from pathlib import Path

import openpyxl
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

OUT = Path(__file__).parent / "test_files"
OUT.mkdir(exist_ok=True)

random.seed(42)

# ─── helpers ────────────────────────────────────────────────────────────────

VENDEDORES = ["Ana García", "Carlos López", "María Rodríguez", "Pedro Martínez", "Sin vendedor"]
CLIENTES = ["Juan Pérez", "Empresa ABC S.A.S.", "María Gómez", "CLIENTE GENERAL", "Restaurante El Buen Sabor"]
FORMAS_PAGO = ["EFECTIVO", "TARJETA", "TRANSFERENCIA", "CREDITO"]

def _fmt_header(ws, row: int = 1) -> None:
    fill = PatternFill("solid", fgColor="0B6B57")
    font = Font(bold=True, color="FFFFFF")
    for cell in ws[row]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center")


# ─── SAZÓN: ventas ──────────────────────────────────────────────────────────

def make_sazon_ventas(months: list[int] = [1, 2, 3], year: int = 2026) -> Path:
    rows = []
    factura = 1001
    for month in months:
        days_in_month = 28 if month == 2 else 30 if month in [4, 6, 9, 11] else 31
        for day in range(1, days_in_month + 1, 1):
            n_sales = random.randint(8, 25)
            for _ in range(n_sales):
                vendedor = random.choice(VENDEDORES)
                cliente = random.choice(CLIENTES)
                subtotal = round(random.uniform(15_000, 180_000), 0)
                iva = round(subtotal * 0.19, 0)
                propina = round(subtotal * random.choice([0, 0, 0.1]), 0)
                descuento = round(subtotal * random.choice([0, 0, 0, 0.05, 0.1]), 0)
                forma = random.choice(FORMAS_PAGO)
                tarjeta = subtotal if forma == "TARJETA" else 0
                efectivo = subtotal if forma == "EFECTIVO" else 0
                transferencia = subtotal if forma == "TRANSFERENCIA" else 0
                credito = subtotal if forma == "CREDITO" else 0

                rows.append({
                    "FECHA": date(year, month, day),
                    "CLIENTE": cliente,
                    "ID CLIENTE": f"CC-{random.randint(1000000, 9999999)}",
                    "NUMERO FACTURA": f"FV-{factura}",
                    "SUBTOTAL": subtotal,
                    "IVA": iva,
                    "TOTAL": subtotal + iva,
                    "DESCUENTO": descuento,
                    "PROPINA": propina,
                    "FORMA DE PAGO": forma,
                    "TARJETA": tarjeta,
                    "EFECTIVO": efectivo,
                    "TRANSFERENCIA": transferencia,
                    "CREDITO": credito,
                    "CORTESIA": descuento,
                    "PROPINA TARJETA": propina if forma == "TARJETA" else 0,
                    "PROPINA RESTO": propina if forma != "TARJETA" else 0,
                    "VALOR FACTURA": subtotal,
                    "VALOR NETO": subtotal - descuento,
                    "ATENDIO": vendedor,
                })
                factura += 1

    df = pd.DataFrame(rows)
    path = OUT / "sazon_ventas_prueba.xlsx"
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="RESUMEN", index=False)
        ws = writer.sheets["RESUMEN"]
        _fmt_header(ws)
        for col in ws.columns:
            ws.column_dimensions[get_column_letter(col[0].column)].width = max(
                len(str(col[0].value or "")), 12
            ) + 2
    print(f"  ✓  {path.name}")
    return path


# ─── SAZÓN: gastos ──────────────────────────────────────────────────────────

PROVEEDORES = [
    "Proveedor Carnes S.A.S.",
    "Distribuidora Verduras Frescas",
    "Arrendamiento local",
    "Nómina y prestaciones",
    "Servicios públicos",
    "Publicidad y marketing",
    "Mantenimiento equipos",
    "Papelería y útiles",
]

def make_sazon_gastos(months: list[int] = [1, 2, 3], year: int = 2026) -> Path:
    MONTH_NAMES = {
        1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril",
        5: "Mayo", 6: "Junio", 7: "Julio", 8: "Agosto",
        9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre",
    }
    path = OUT / "sazon_gastos_prueba.xlsx"
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    for month in months:
        ws = wb.create_sheet(MONTH_NAMES[month])
        ws.append(["PROVEEDOR / CONCEPTO", "VALOR GASTO"])
        ws["A1"].font = Font(bold=True)
        ws["B1"].font = Font(bold=True)
        for prov in PROVEEDORES:
            valor = round(random.uniform(500_000, 8_000_000), 0)
            ws.append([prov, valor])
        ws.column_dimensions["A"].width = 36
        ws.column_dimensions["B"].width = 18

    wb.save(path)
    print(f"  ✓  {path.name}")
    return path


# ─── TLG: balance de prueba ─────────────────────────────────────────────────

CUENTAS_TLG = [
    # (codigo, nombre, clase, transaccional, saldo_final)
    ("1105", "Caja general", "1", "SI", 12_500_000),
    ("1110", "Bancos cuenta corriente", "1", "SI", 45_300_000),
    ("1305", "Clientes nacionales", "1", "SI", 28_700_000),
    ("1330", "Anticipos y avances", "1", "SI", 5_200_000),
    ("1520", "Construcciones y edificaciones", "1", "NO", 120_000_000),
    ("1524", "Depreciación acumulada edificaciones", "1", "NO", -15_000_000),
    ("1540", "Maquinaria y equipo", "1", "NO", 55_000_000),
    ("1592", "Depreciación acumulada maquinaria", "1", "NO", -8_000_000),
    ("2105", "Bancos nacionales", "2", "SI", -35_200_000),
    ("2205", "Proveedores nacionales", "2", "SI", -22_400_000),
    ("2365", "Retención en la fuente por pagar", "2", "NO", -1_800_000),
    ("3005", "Capital suscrito y pagado", "3", "NO", -120_000_000),
    ("3605", "Utilidad del ejercicio", "3", "NO", -64_300_000),
    ("4135", "Comercio al por mayor y al por menor", "4", "NO", -95_000_000),
    ("4175", "Servicios de restaurante", "4", "NO", -38_500_000),
    ("5105", "Gastos de personal", "5", "NO", 42_000_000),
    ("5110", "Honorarios", "5", "NO", 8_400_000),
    ("5115", "Impuestos", "5", "NO", 3_200_000),
    ("5120", "Arrendamientos", "5", "NO", 9_600_000),
    ("5195", "Diversos gastos administrativos", "5", "NO", 5_100_000),
    ("6135", "Compras de mercancía", "6", "NO", 55_200_000),
]

TERCEROS = [
    ("900123456-1", "THE LATAM GROUP S.A.S."),
    ("800234567-2", "Banco de Bogotá"),
    ("901345678-3", "Cliente TLG 1"),
]

def make_tlg_balance(month: int = 3, year: int = 2026) -> Path:
    MONTH_NAMES = {3: "Marzo", 12: "Diciembre", 1: "Enero", 6: "Junio"}
    month_name = MONTH_NAMES.get(month, f"Mes{month:02d}")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Balance de prueba"

    # Header info (metadata that the cleaner looks for)
    ws["A1"] = "THE LATAM GROUP S.A.S."
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"NIT: 900.123.456-1"
    ws["A3"] = f"Balance de Prueba - Del Enero {year} al {month_name} {year}"
    ws["A3"].font = Font(italic=True)
    ws.append([])  # blank row

    # Column headers (row 5)
    headers = [
        "NIVEL", "TRANSACCIONAL", "CODIGO CUENTA CONTABLE", "NOMBRE CUENTA CONTABLE",
        "IDENTIFICACION", "SUCURSAL", "NOMBRE TERCERO",
        "SALDO INICIAL", "MOVIMIENTO DEBITO", "MOVIMIENTO CREDITO", "SALDO FINAL",
    ]
    ws.append(headers)
    header_row = ws.max_row
    for cell in ws[header_row]:
        cell.fill = PatternFill("solid", fgColor="1F2937")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for codigo, nombre, clase, trans, saldo in CUENTAS_TLG:
        nivel = "AUXILIAR" if trans == "SI" else "CUENTA"
        saldo_ini = round(saldo * 0.85, 0)
        if saldo >= 0:
            debito = round(saldo * 0.4, 0)
            credito = round(saldo * 0.4 - (saldo - saldo_ini), 0)
        else:
            credito = round(abs(saldo) * 0.4, 0)
            debito = round(abs(saldo) * 0.4 - (abs(saldo) - abs(saldo_ini)), 0)
        tercero_id, tercero_nombre = random.choice(TERCEROS) if trans == "SI" else ("", "")
        ws.append([
            nivel, trans, codigo, nombre,
            tercero_id, "001", tercero_nombre,
            saldo_ini, debito, credito, saldo,
        ])

    # Column widths
    widths = [12, 14, 24, 45, 18, 8, 35, 18, 18, 18, 18]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    path = OUT / f"tlg_balance_prueba_{month_name.lower()}_{year}.xlsx"
    wb.save(path)
    print(f"  ✓  {path.name}")
    return path


# ─── MENSUALIZADOS: varios balances ─────────────────────────────────────────

def make_mensualizados_balances(months: list[int] = [1, 2, 3], year: int = 2026) -> list[Path]:
    paths = []
    for month in months:
        p = make_tlg_balance(month=month, year=year)
        new_name = OUT / f"mensualizado_balance_{month:02d}_{year}.xlsx"
        p.rename(new_name)
        paths.append(new_name)
        print(f"    → renombrado a {new_name.name}")
    return paths


# ─── main ────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    print("\n=== Generando archivos de prueba ===\n")

    print("[ Sazón ]")
    make_sazon_ventas(months=[1, 2, 3], year=2026)
    make_sazon_gastos(months=[1, 2, 3], year=2026)

    print("\n[ TLG ]")
    p = make_tlg_balance(month=3, year=2026)
    p.rename(OUT / "tlg_balance_prueba_marzo_2026.xlsx")
    print(f"    → tlg_balance_prueba_marzo_2026.xlsx")

    print("\n[ Mensualizados ]")
    for month in [1, 2, 3]:
        make_tlg_balance(month=month, year=2026)
        src = OUT / f"tlg_balance_prueba_{'marzo' if month==3 else 'enero' if month==1 else 'febrero'}_{2026}.xlsx"
        dst = OUT / f"mensualizado_balance_{month:02d}_2026.xlsx"
        if src.exists() and not dst.exists():
            src.rename(dst)

    print(f"\n✅  Archivos generados en: {OUT}\n")
    for f in sorted(OUT.glob("*.xlsx")):
        size_kb = f.stat().st_size // 1024
        print(f"   {f.name:55s}  {size_kb:>4} KB")
