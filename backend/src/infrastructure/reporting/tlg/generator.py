from __future__ import annotations

from io import BytesIO

import pandas as pd
from openpyxl import Workbook, load_workbook
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer


def _month_sheet_name(month: str | None, year: str | None) -> str:
    if not month:
        return "Bp periodo"
    yy = str(year or "")[-2:] or "26"
    return f"Bp {month[:3].title()} {yy}"


def update_tlg_financial_statements(template_file, detail: pd.DataFrame, metadata: dict[str, str | None]) -> bytes:
    if template_file is not None:
        template_file.seek(0)
        wb = load_workbook(template_file)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Resumen TLG"
        ws["A1"] = "Estados Financieros TLG - archivo generado"

    sheet_name = _month_sheet_name(metadata.get("mes"), metadata.get("anio"))
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        ws.delete_rows(1, ws.max_row)
    else:
        ws = wb.create_sheet(sheet_name)

    headers = detail.columns.tolist()
    ws.append(headers)
    for row in detail.itertuples(index=False):
        ws.append(list(row))

    if "Resumen generado" in wb.sheetnames:
        summary_ws = wb["Resumen generado"]
        summary_ws.delete_rows(1, summary_ws.max_row)
    else:
        summary_ws = wb.create_sheet("Resumen generado")
    summary_ws.append(["Campo", "Valor"])
    for key, value in metadata.items():
        summary_ws.append([key, value])

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


def build_tlg_summary_excel(summary: dict[str, object], metadata: dict[str, str | None]) -> bytes:
    output = BytesIO()
    with pd.ExcelWriter(output, engine="xlsxwriter") as writer:
        pd.DataFrame([metadata]).to_excel(writer, sheet_name="Validacion", index=False)
        pd.DataFrame([summary["metrics"]]).to_excel(writer, sheet_name="Indicadores", index=False)
        summary["balance"].to_excel(writer, sheet_name="ESF", index=False)
        summary["income_statement"].to_excel(writer, sheet_name="Estado Resultados", index=False)
        summary["detail"].to_excel(writer, sheet_name="Balance limpio", index=False)
    return output.getvalue()


def build_tlg_management_pdf(text: str, metrics: dict[str, float], metadata: dict[str, str | None]) -> bytes:
    output = BytesIO()
    doc = SimpleDocTemplate(output, pagesize=letter, rightMargin=42, leftMargin=42, topMargin=42, bottomMargin=42)
    styles = getSampleStyleSheet()
    story = [
        Paragraph("Resumen gerencial TLG", styles["Title"]),
        Paragraph(f"Periodo: {metadata.get('periodo') or 'No detectado'}", styles["Normal"]),
        Spacer(1, 12),
        Paragraph(text, styles["BodyText"]),
        Spacer(1, 12),
        Paragraph(f"Diferencia de cuadre: ${metrics.get('diferencia_cuadre', 0):,.0f}".replace(",", "."), styles["Normal"]),
    ]
    doc.build(story)
    output.seek(0)
    return output.getvalue()
